from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from datetime import datetime
import subprocess

class ExcelManager:
    def __init__(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.setup_styles()

    def setup_styles(self):
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))

        # colorCode
        self.custom_font = Font(name='Calibri', size=10,
                                bold=True, italic=True, color='000000')

        self.custom_alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=False)

        self.fill_colors = {
            'green': PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),
            'yellow': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
            'red': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            'orange': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        }

    def conv2excel(self, dataframe, current_lesson, lesson):
        self.lesson = lesson
        for row in dataframe_to_rows(dataframe, index=False, header=True):
            self.sheet.append(row)
        self.adjust_column_widths()
        self.format_cells(current_lesson)

    def adjust_column_widths(self):
        self.sheet.column_dimensions['B'].width = 30
        self.sheet.column_dimensions['A'].width = 5
        for col in range(67, 91):  # ASCII values for 'C' to 'Z'
            self.sheet.column_dimensions[chr(col)].width = 11

    def format_cells(self, current_lesson):
        for i, row in enumerate(self.sheet.iter_rows()):
            for j, cell in enumerate(row):
                self.apply_cell_style(cell, i, j)

        # Highlight the current lesson column
        self.sheet.cell(row=1, column=current_lesson +
                        2).fill = self.fill_colors['orange']

    def apply_cell_style(self, cell, i, j):
        cell.border = self.thin_border
        cell.font = self.custom_font
        cell.alignment = self.custom_alignment
        if cell.value != "" and i > 0 and j > 1 and self.lesson[j - 2]:
            ac, wa = map(lambda x: int(x.split(":")[1]), str(
                cell.value).split("  "))
            cell.fill = self.determine_fill_color(ac, wa, self.lesson[j - 2])


    def determine_fill_color(self, ac, wa, ls_j):
        ratio = ac / ls_j
        if ratio == 1:
            return self.fill_colors["green"]
        elif ratio >= 0.5:
            return self.fill_colors["yellow"]
        else:
            return self.fill_colors["red"]

    def save(self, excel_path,clas):
        excel_path += f"{clas}{datetime.today().strftime("(%d-%m-%Y-%H-%M-%S)")}.xlsx"
        self.workbook.save(excel_path)
        return excel_path

    def save_and_open(self, excel_path,clas):
        excel_path += f"{clas}{datetime.today().strftime("(%d-%m-%Y-%H-%M-%S)")}.xlsx"
        self.workbook.save(excel_path)
        subprocess.Popen(["start", excel_path],shell = True)